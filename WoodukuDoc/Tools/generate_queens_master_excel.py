import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


PALETTE = [
    "F4B183",
    "FFD966",
    "A9D18E",
    "8DD3C7",
    "9DC3E6",
    "B4C6E7",
    "C9B1FF",
    "D9A7C7",
    "F4CCCC",
    "D5C4A1",
    "B7B7B7",
    "C6E0B4",
]

DIFFICULTY_NAMES = {
    0: "普通",
    1: "困难",
    2: "极难",
}


def parse_args():
    parser = argparse.ArgumentParser(description="Generate an Excel atlas for Queens Master levels.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument(
        "--hide-solution",
        action="store_true",
        help="Only show pre-placed queens instead of the full solution.",
    )
    parser.add_argument("--verify", action="store_true")
    return parser.parse_args()


def board_anchor(level_index, size, boards_per_row):
    tile_row = level_index // boards_per_row
    tile_col = level_index % boards_per_row
    row = tile_row * (size + 4) + 1
    col = tile_col * (size + 1) + 1
    return row, col


def create_readme(workbook, total_levels, show_solution):
    sheet = workbook.active
    sheet.title = "说明"
    rows = [
        ("Queens Master 关卡图册", ""),
        ("关卡总数", total_levels),
        ("显示模式", "区域 + 完整答案" if show_solution else "区域 + 预置皇后"),
        ("♛", "答案皇后"),
        ("★", "开局预置且不可移除的皇后"),
        ("关卡索引", "可筛选，并可点击“图案位置”跳转到对应棋盘"),
        ("尺寸工作表", "相同尺寸的关卡按从左到右、从上到下排列"),
    ]
    for row_index, values in enumerate(rows, 1):
        sheet.cell(row_index, 1, values[0])
        sheet.cell(row_index, 2, values[1])

    sheet["A1"].font = Font(size=16, bold=True)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 60
    sheet.freeze_panes = "A2"


def create_index_sheet(workbook):
    sheet = workbook.create_sheet("关卡索引")
    headers = [
        "关卡ID",
        "来源名",
        "尺寸",
        "难度",
        "难度分",
        "预置皇后数",
        "资源包",
        "图案位置",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    widths = [10, 42, 9, 10, 12, 14, 12, 14]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    sheet.freeze_panes = "A2"
    return sheet


def create_size_sheet(workbook, size):
    sheet = workbook.create_sheet(f"{size}x{size}")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A1"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    return sheet


def write_level_tile(
    sheet,
    level,
    level_id,
    tile_index,
    boards_per_row,
    show_solution,
    thin_border,
    center,
):
    size = level["size"]
    start_row, start_col = board_anchor(tile_index, size, boards_per_row)
    end_col = start_col + size - 1

    title = (
        f"#{level_id:04d}  "
        f"{DIFFICULTY_NAMES.get(level.get('difficulty', 0), '未知')}  "
        f"难度分 {level.get('difficultyScore', 0)}"
    )
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=end_col,
    )
    title_cell = sheet.cell(start_row, start_col, title)
    title_cell.font = Font(bold=True, size=9, color="0563C1", underline="single")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
    title_cell.hyperlink = Hyperlink(
        ref=title_cell.coordinate,
        location=f"'关卡索引'!A{level_id + 1}",
        display=title,
        tooltip="返回关卡索引",
    )

    solution_cols = level.get("solutionCols") or []
    fixed_positions = {
        (cell["r"], cell["c"])
        for cell in level.get("fixedQueenCells") or []
    }

    for row in range(size):
        sheet.row_dimensions[start_row + 1 + row].height = 20
        for col in range(size):
            cell = sheet.cell(start_row + 1 + row, start_col + col)
            region = level["regions"][row * size + col]
            cell.fill = PatternFill("solid", fgColor=PALETTE[region % len(PALETTE)])
            cell.border = thin_border
            cell.alignment = center

            if (row, col) in fixed_positions:
                cell.value = "★"
                cell.font = Font(bold=True, size=12, color="9C0006")
            elif show_solution and row < len(solution_cols) and solution_cols[row] == col:
                cell.value = "♛"
                cell.font = Font(bold=True, size=12, color="000000")

    source_row = start_row + size + 1
    sheet.merge_cells(
        start_row=source_row,
        start_column=start_col,
        end_row=source_row,
        end_column=end_col,
    )
    source_cell = sheet.cell(source_row, start_col, level.get("sourceName", ""))
    source_cell.font = Font(size=7, color="666666")
    source_cell.alignment = Alignment(horizontal="center", shrink_to_fit=True)

    return sheet.title, start_row, start_col


def generate_workbook(payload, output_path, show_solution):
    levels = payload["levels"]
    grouped = defaultdict(list)
    for level_id, level in enumerate(levels, 1):
        grouped[level["size"]].append((level_id, level))

    workbook = Workbook()
    create_readme(workbook, len(levels), show_solution)
    index_sheet = create_index_sheet(workbook)

    thin_side = Side(style="thin", color="666666")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center")
    index_rows = {}

    for size in sorted(grouped):
        sheet = create_size_sheet(workbook, size)
        boards_per_row = 5 if size <= 6 else 4 if size <= 9 else 3
        for column in range(1, boards_per_row * (size + 1) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 3.2

        for tile_index, (level_id, level) in enumerate(grouped[size]):
            sheet_name, row, col = write_level_tile(
                sheet,
                level,
                level_id,
                tile_index,
                boards_per_row,
                show_solution,
                thin_border,
                center,
            )
            index_rows[level_id] = (sheet_name, row, col)

    for level_id, level in enumerate(levels, 1):
        sheet_name, row, col = index_rows[level_id]
        fixed_count = len(level.get("fixedQueenCells") or [])
        values = [
            level_id,
            level.get("sourceName", ""),
            f"{level['size']}x{level['size']}",
            DIFFICULTY_NAMES.get(level.get("difficulty", 0), "未知"),
            level.get("difficultyScore", 0),
            fixed_count,
            (level_id - 1) // 256 + 1,
        ]
        excel_row = level_id + 1
        for column, value in enumerate(values, 1):
            index_sheet.cell(excel_row, column, value)

        anchor = f"{get_column_letter(col)}{row}"
        link_cell = index_sheet.cell(excel_row, 8, "查看图案")
        link_cell.hyperlink = Hyperlink(
            ref=link_cell.coordinate,
            location=f"'{sheet_name}'!{anchor}",
            display="查看图案",
            tooltip=f"跳转到 {sheet_name}!{anchor}",
        )
        link_cell.style = "Hyperlink"

    index_sheet.auto_filter.ref = f"A1:H{len(levels) + 1}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def verify_workbook(output_path, expected_levels, expected_sizes):
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        index_sheet = workbook["关卡索引"]
        indexed_levels = index_sheet.max_row - 1
        actual_sizes = {
            int(name.split("x", 1)[0])
            for name in workbook.sheetnames
            if "x" in name and name.split("x", 1)[0].isdigit()
        }
        if indexed_levels != expected_levels:
            raise RuntimeError(
                f"Index count mismatch: expected={expected_levels}, actual={indexed_levels}"
            )
        if actual_sizes != expected_sizes:
            raise RuntimeError(
                f"Size sheet mismatch: expected={sorted(expected_sizes)}, "
                f"actual={sorted(actual_sizes)}"
            )
    finally:
        workbook.close()


def main():
    args = parse_args()
    payload = json.loads(args.input.read_text(encoding="utf-8"))
    levels = payload.get("levels") or []
    if not levels:
        raise RuntimeError("No levels found in input file.")

    show_solution = not args.hide_solution
    generate_workbook(payload, args.output, show_solution)

    if args.verify:
        verify_workbook(
            args.output,
            expected_levels=len(levels),
            expected_sizes={level["size"] for level in levels},
        )

    print(f"levels={len(levels)}")
    print(f"sizes={','.join(str(size) for size in sorted({level['size'] for level in levels}))}")
    print(f"show_solution={show_solution}")
    print(f"output={args.output}")
    print(f"bytes={args.output.stat().st_size}")


if __name__ == "__main__":
    main()
